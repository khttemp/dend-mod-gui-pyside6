import os
import struct
import traceback
import openpyxl
import configparser

import program.sub.textSetting as textSetting
from program.sub.encodingClass import SJISEncodingObject
from program.sub.errorLogClass import ErrorLogObj


class ExcelWidget:
    def __init__(self, filePath, decryptFile, configPath):
        self.encObj = SJISEncodingObject()
        self.errObj = ErrorLogObj()
        self.filePath = filePath
        self.decryptFile = decryptFile
        self.configPath = configPath
        self.modelNameMode = 0
        self.flagHexMode = 0
        self.MODEL_NAME = 0
        self.HEX_FLAG = 1
        self.row = -1
        self.errorLogList = []
        self.errorMessage = ""
        self.warningLogList = []
        self.warningMessage = ""

    def printError(self, error):
        self.errObj.write(error)

    def extractExcel(self):
        self.errorLogList = []
        self.errorMessage = ""
        self.warningLogList = []
        self.warningMessage = ""

        wb = openpyxl.Workbook()

        # シート初期化
        defSheetNameList = wb.sheetnames
        for sheetName in defSheetNameList:
            wb.remove(wb[sheetName])

        # TabList
        tabList = textSetting.textList["railEditor"]["railComboValue"]
        try:
            for index, tabName in enumerate(tabList):
                wb.create_sheet(index=index, title=tabName)
                self.extractRailDataInfo(index, wb[tabName])

            wb.save(self.filePath)
            return True
        except PermissionError:
            self.errorMessage = textSetting.textList["errorList"]["E94"]
            return False
        except Exception:
            self.printError(traceback.format_exc())
            self.errorMessage = textSetting.textList["errorList"]["E14"]
            return False

    def extractRailDataInfo(self, sheetIndex, ws):
        configRead = configparser.ConfigParser()
        configRead.read(self.configPath, encoding="utf-8")
        self.modelNameMode = int(configRead.get("MODEL_NAME_MODE", "mode"))
        self.flagHexMode = int(configRead.get("FLAG_MODE", "mode"))

        self.row = 1
        # BGM、配置情報
        if sheetIndex == 0:
            self.extractBgmAndPosInfo(ws)
        # 要素１
        elif sheetIndex == 1:
            self.extractElse1Info(ws)
        # smf情報
        elif sheetIndex == 2:
            self.extractSmfInfo(ws)
        # 駅名位置情報
        elif sheetIndex == 3:
            self.extractStationInfo(ws)
        # 要素２
        elif sheetIndex == 4:
            self.extractElse2Info(ws)
        # CPU情報
        elif sheetIndex == 5:
            self.extractCpuInfo(ws)
        # コミックスクリプト
        elif sheetIndex == 6:
            self.extractComicScriptInfo(ws)
        # レール情報
        elif sheetIndex == 7:
            self.extractRailInfo(ws)
        # 要素３
        elif sheetIndex == 8:
            self.extractElse3Info(ws)
        # 要素４
        elif sheetIndex == 9:
            self.extractElse4Info(ws)
        # AMB情報
        elif sheetIndex == 10:
            self.extractAmbInfo(ws)

    def extractBgmAndPosInfo(self, ws):
        # ver
        ws.cell(self.row, 1).value = self.decryptFile.ver
        self.row += 2

        # BGM
        ws.cell(self.row, 1).value = "BGM"
        ws.cell(self.row, 2).value = self.decryptFile.musicCnt
        self.row += 1
        for musicInfo in self.decryptFile.musicList:
            for idx, music in enumerate(musicInfo):
                ws.cell(self.row, 1 + idx).value = music
            self.row += 1
        self.row += 1

        # 車両の初期レール位置
        ws.cell(self.row, 1).value = "RailPos"
        ws.cell(self.row, 2).value = self.decryptFile.trainCnt
        self.row += 1
        for trainInfo in self.decryptFile.trainList:
            for idx, train in enumerate(trainInfo):
                ws.cell(self.row, 1 + idx).value = train
            self.row += 1
        self.row += 1

        # ダミー位置？
        ws.cell(self.row, 1).value = "RailPos2"
        self.row += 1
        for trainInfo2 in self.decryptFile.trainList2:
            for idx, train2 in enumerate(trainInfo2):
                ws.cell(self.row, 1 + idx).value = train2
            self.row += 1
        self.row += 1

        # 試運転、二人バトルの初期レール位置
        ws.cell(self.row, 1).value = "FreeRunOrVSPos"
        self.row += 1
        for trainInfo3 in self.decryptFile.trainList3:
            for idx, train3 in enumerate(trainInfo3):
                ws.cell(self.row, 1 + idx).value = train3
            self.row += 1
        self.row += 1

        # 駅表示を始める番号
        ws.cell(self.row, 1).value = "stationNo"
        ws.cell(self.row, 2).value = self.decryptFile.stationNo
        self.row += 2

        ws.cell(self.row, 1).value = "RailPos4"
        self.row += 1
        for trainInfo4 in self.decryptFile.trainList4:
            for idx, train4 in enumerate(trainInfo4):
                ws.cell(self.row, 1 + idx).value = train4
            self.row += 1
        self.row += 1

        ws.cell(self.row, 1).value = "stationNo2"
        ws.cell(self.row, 2).value = self.decryptFile.stationNo2
        self.row += 2

        # レール名
        ws.cell(self.row, 1).value = "railName"
        ws.cell(self.row, 2).value = self.decryptFile.railStationName

    def extractElse1Info(self, ws):
        ws.cell(self.row, 1).value = "else1-1"
        ws.cell(self.row, 2).value = self.decryptFile.else1List[0]
        self.row += 2

        ws.cell(self.row, 1).value = "else1-2"
        ws.cell(self.row, 2).value = len(self.decryptFile.else1List[1:])
        self.row += 1

        for else1Info in self.decryptFile.else1List[1:]:
            for idx, else1 in enumerate(else1Info):
                ws.cell(self.row, 1 + idx).value = else1
            self.row += 1
        self.row += 1

        # light情報
        ws.cell(self.row, 1).value = "light"
        ws.cell(self.row, 2).value = len(self.decryptFile.lightList)
        self.row += 1

        for lightName in self.decryptFile.lightList:
            ws.cell(self.row, 1).value = lightName
            self.row += 1
        self.row += 1
        
        # base bin
        ws.cell(self.row, 1).value = "baseBin"
        ws.cell(self.row, 2).value = len(self.decryptFile.baseBinList)
        self.row += 1

        for baseBinName in self.decryptFile.baseBinList:
            ws.cell(self.row, 1).value = baseBinName
            self.row += 1
        self.row += 1
        
        ws.cell(self.row, 1).value = "binAnime"
        ws.cell(self.row, 2).value = len(self.decryptFile.binAnimeList)
        self.row += 1

        for binAnimeInfo in self.decryptFile.binAnimeList:
            for idx, binAnime in enumerate(binAnimeInfo):
                ws.cell(self.row, 1 + idx).value = binAnime
            self.row += 1

    def extractSmfInfo(self, ws):
        ws.cell(self.row, 1).value = "MdlCnt"
        ws.cell(self.row, 2).value = len(self.decryptFile.smfList)
        self.row += 1

        for smfIdx, smfInfo in enumerate(self.decryptFile.smfList):
            ws.cell(self.row, 1).value = smfIdx
            idx = 0
            for smf in smfInfo[:-1]:
                ws.cell(self.row, 2 + idx).value = smf
                idx += 1
            ws.cell(self.row, 2 + idx).value = len(smfInfo[-1])
            idx += 1
            animeIdx = idx
            for smfAnimeInfo in smfInfo[-1]:
                idx = animeIdx
                for smfAnime in smfAnimeInfo:
                    ws.cell(self.row, 2 + idx).value = smfAnime
                    idx += 1
                self.row += 1
            if len(smfInfo[-1]) == 0:
                self.row += 1

    def extractStationInfo(self, ws):
        ws.cell(self.row, 1).value = "STCnt"
        ws.cell(self.row, 2).value = len(self.decryptFile.stationNameList)
        self.row += 1

        for stIdx, stationNameInfo in enumerate(self.decryptFile.stationNameList):
            ws.cell(self.row, 1).value = stIdx
            for idx, stInfo in enumerate(stationNameInfo):
                ws.cell(self.row, 2 + idx).value = stInfo
            self.row += 1

    def extractElse2Info(self, ws):
        ws.cell(self.row, 1).value = "else2"
        ws.cell(self.row, 2).value = len(self.decryptFile.else2List)
        self.row += 1

        for else2Info in self.decryptFile.else2List:
            for idx, else2 in enumerate(else2Info):
                ws.cell(self.row, 1 + idx).value = else2
            self.row += 1

    def extractCpuInfo(self, ws):
        ws.cell(self.row, 1).value = "CPU"
        ws.cell(self.row, 2).value = len(self.decryptFile.cpuList)
        self.row += 1

        for cpuIdx, cpuInfo in enumerate(self.decryptFile.cpuList):
            ws.cell(self.row, 1).value = cpuIdx
            for idx, cpu in enumerate(cpuInfo):
                ws.cell(self.row, 2 + idx).value = cpu
            self.row += 1

    def extractComicScriptInfo(self, ws):
        ws.cell(self.row, 1).value = "ComicScript"
        ws.cell(self.row, 2).value = len(self.decryptFile.comicScriptList)
        self.row += 1

        for scriptIdx, comicScriptInfo in enumerate(self.decryptFile.comicScriptList):
            ws.cell(self.row, 1).value = scriptIdx
            for idx, comicScript in enumerate(comicScriptInfo):
                ws.cell(self.row, 2 + idx).value = comicScript
            self.row += 1

    def extractRailInfo(self, ws):
        ws.cell(self.row, 1).value = "RailCnt"
        ws.cell(self.row, 2).value = len(self.decryptFile.railList)
        self.row += 2
        mdlList = [x[0] for x in self.decryptFile.smfList]

        titleList = [
            "index",
            "prev_rail",
            "block",
            "dir_x",
            "dir_y",
            "dir_z",
            "mdl_no",
            "mdl_kasen",
            "mdl_kasenchu",
            "per",
            "flg",
            "flg",
            "flg",
            "flg",
            "rail_data",
            "next_rail",
            "next_no",
            "prev_rail",
            "prev_no",
        ]

        for idx, title in enumerate(titleList):
            ws.cell(self.row, 1 + idx).value = title
        self.row += 1

        for railInfo in self.decryptFile.railList:
            # index
            ws.cell(self.row, 1).value = railInfo[0]

            # prev_rail
            ws.cell(self.row, 2).value = railInfo[1]

            # block
            ws.cell(self.row, 3).value = railInfo[2]

            # dir
            for i in range(3):
                ws.cell(self.row, 4 + i).value = railInfo[3 + i]

            # mdl_no
            mdl_no = railInfo[6]
            if self.modelNameMode == self.MODEL_NAME:
                mdl_no = self.getSmfModelName(mdl_no, mdlList)
            ws.cell(self.row, 7).value = mdl_no

            # kasen
            kasen = railInfo[7]
            if self.modelNameMode == self.MODEL_NAME:
                kasen = self.getSmfModelName(kasen, mdlList)
            ws.cell(self.row, 8).value = kasen

            # kasenchu
            kasenchu = railInfo[8]
            if self.modelNameMode == self.MODEL_NAME:
                kasenchu = self.getSmfModelName(kasenchu, mdlList)
            ws.cell(self.row, 9).value = kasenchu

            # per
            ws.cell(self.row, 10).value = railInfo[9]

            # flg
            for i in range(4):
                flg = railInfo[10 + i]
                if self.flagHexMode == self.HEX_FLAG:
                    flg = self.toHex(flg)
                ws.cell(self.row, 11 + i).value = flg

            # raildata
            raildata = railInfo[14]
            ws.cell(self.row, 15).value = raildata

            for i in range(raildata):
                for j in range(4):
                    ws.cell(self.row, 16 + 4*i + j).value = railInfo[15 + 4*i + j]
            self.row += 1

    def extractElse3Info(self, ws):
        ws.cell(self.row, 1).value = "else3"
        ws.cell(self.row, 2).value = len(self.decryptFile.else3List)
        self.row += 1

        for else3Idx, else3Info in enumerate(self.decryptFile.else3List):
            ws.cell(self.row, 1).value = else3Idx
            idx = 0
            for else3 in else3Info[:-1]:
                ws.cell(self.row, 2 + idx).value = else3
                idx += 1
            ws.cell(self.row, 2 + idx).value = len(else3Info[-1])
            idx += 1
            listIdx = idx
            for listInfo in else3Info[-1]:
                idx = listIdx
                for e3list in listInfo:
                    ws.cell(self.row, 2 + idx).value = e3list
                    idx += 1
                self.row += 1
            if len(else3Info[-1]) == 0:
                self.row += 1

    def extractElse4Info(self, ws):
        ws.cell(self.row, 1).value = "else4"
        ws.cell(self.row, 2).value = len(self.decryptFile.else4List)
        self.row += 1

        for else4Idx, else4Info in enumerate(self.decryptFile.else4List):
            ws.cell(self.row, 1).value = else4Idx
            for idx, else4 in enumerate(else4Info):
                ws.cell(self.row, 2 + idx).value = else4
            self.row += 1

    def extractAmbInfo(self, ws):
        ws.cell(self.row, 1).value = "AmbCnt"
        ws.cell(self.row, 2).value = len(self.decryptFile.ambList)
        self.row += 2
        mdlList = [x[0] for x in self.decryptFile.smfList]

        titleList = [
            "index",
            "rail_no",
            "priority",
            "fog",
            "mdl_no",
            "rail_pos",
            "pos_x",
            "pos_y",
            "pos_z",
            "rot_x",
            "rot_y",
            "rot_z",
            "per"
        ]
        for idx, title in enumerate(titleList):
            ws.cell(self.row, 1 + idx).value = title
        self.row += 1

        for ambIdx, ambInfo in enumerate(self.decryptFile.ambList):
            ws.cell(self.row, 1).value = ambIdx
            for idx, amb in enumerate(ambInfo):
                # mdl_no
                if idx == 3:
                    if self.modelNameMode == self.MODEL_NAME:
                        amb = self.getSmfModelName(amb, mdlList)
                ws.cell(self.row, 2 + idx).value = amb
            self.row += 1

    def toHex(self, num):
        return "0x{:02x}".format(num)

    def getSmfModelName(self, modelIndex, smfNameList):
        if modelIndex < 0 or modelIndex >= len(smfNameList):
            return modelIndex
        modelName = smfNameList[modelIndex]
        if self.isModelNameDup(modelName, smfNameList):
            return modelIndex
        else:
            return modelName

    def isModelNameDup(self, modelName, smfNameList):
        if type(modelName) is str:
            modelNameList = [x[0] for x in smfNameList if x == modelName]
            if len(modelNameList) > 1:
                return True
            return False
        return False

    def loadExcelData(self):
        wb = openpyxl.load_workbook(self.filePath, data_only=True)
        configRead = configparser.ConfigParser()
        configRead.read(self.configPath, encoding="utf-8")
        self.modelNameMode = int(configRead.get("MODEL_NAME_MODE", "mode"))
        self.flagHexMode = int(configRead.get("FLAG_MODE", "mode"))

        self.newByteArr = bytearray()
        self.newSmfList = []
        self.else4Dict = {}
        self.ambDict = {}
        self.else3Dict = {}
        self.excelSheet = ""
        self.excelCell = None
        # TabList
        tabList = textSetting.textList["railEditor"]["railComboValue"]
        try:
            for tabName in tabList:
                if tabName not in wb.sheetnames:
                    return (False, {"message":textSetting.textList["errorList"]["E95"].format(tabName) })
            # BGM、配置情報
            self.getBgmAndPosInfo(wb[tabList[0]])
            # 要素１
            self.getElse1Info(wb[tabList[1]])
            # smf情報
            self.getSmfInfo(wb[tabList[2]])
            # 駅名位置情報
            self.getStationInfo(wb[tabList[3]])
            # 要素２
            self.getElse2Info(wb[tabList[4]])
            # CPU情報
            self.getCpuInfo(wb[tabList[5]])
            # コミックスクリプト
            self.getComicScriptInfo(wb[tabList[6]])
            # 要素４
            self.getElse4Info(wb[tabList[9]])
            # AMB情報
            self.getAmbInfo(wb[tabList[10]])
            # 要素３
            self.getElse3Info(wb[tabList[8]])
            # レール情報
            self.getRailInfo(wb[tabList[7]])

            if len(self.errorLogList) > 0:
                dirPath = os.path.dirname(self.filePath)
                errPath = os.path.join(dirPath, "railError.log")
                w = open(errPath, "w", encoding="utf-8")
                for err in self.errorLogList:
                    w.write(err + "\n")
                w.close()
                self.errorMessage = textSetting.textList["errorList"]["E118"].format("railError.log")
                return (False, {"message":self.errorMessage})

            return (True, {"message":textSetting.textList["infoList"]["I117"], "data":self.newByteArr})
        except Exception:
            self.printError(traceback.format_exc())
            return (False, {"message":textSetting.textList["errorList"]["E135"].format(self.excelSheet, self.excelCell.coordinate)})

    def getBgmAndPosInfo(self, ws):
        self.excelSheet = ws.title
        # ver
        self.excelCell = ws.cell(1, 1)
        ver = self.excelCell.value
        if ver is None:
            self.errorLogList.append(textSetting.textList["errorList"]["E99"])
            return
        bVer = self.encObj.convertByteArray(ver)
        self.newByteArr.extend(bVer)

        # レール名
        search = "railName"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        railStationName = self.excelCell.value
        bRailStationName = self.encObj.convertByteArray(railStationName)
        self.newByteArr.append(len(bRailStationName))
        self.newByteArr.extend(bRailStationName)

        # BGM
        search = "BGM"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        musicCnt = self.excelCell.value
        self.newByteArr.append(musicCnt)
        row += 1
        for i in range(musicCnt):
            self.excelCell = ws.cell(row, 1)
            musicFile = self.excelCell.value
            bMusicFile = self.encObj.convertByteArray(musicFile)
            self.newByteArr.append(len(bMusicFile))
            self.newByteArr.extend(bMusicFile)

            self.excelCell = ws.cell(row, 2)
            musicName = self.excelCell.value
            bMusicName = self.encObj.convertByteArray(musicName)
            self.newByteArr.append(len(bMusicName))
            self.newByteArr.extend(bMusicName)

            self.excelCell = ws.cell(row, 3)
            start = self.excelCell.value
            self.newByteArr.extend(struct.pack("<f", start))
            self.excelCell = ws.cell(row, 4)
            loopStart = self.excelCell.value
            self.newByteArr.extend(struct.pack("<f", loopStart))
            row += 1

        # 車両の初期レール位置
        search = "RailPos"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        trainCnt = self.excelCell.value
        self.newByteArr.append(trainCnt)
        row += 1
        
        for i in range(trainCnt):
            self.excelCell = ws.cell(row, 1)
            railNo = self.excelCell.value
            hRailNo = struct.pack("<h", railNo)
            self.newByteArr.extend(hRailNo)

            self.excelCell = ws.cell(row, 2)
            railPos = self.excelCell.value
            hRailPos = struct.pack("<h", railPos)
            self.newByteArr.extend(hRailPos)

            self.excelCell = ws.cell(row, 3)
            self.newByteArr.append(self.excelCell.value)

            self.excelCell = ws.cell(row, 4)
            f1 = self.excelCell.value
            tempF = struct.pack("<f", f1)
            self.newByteArr.extend(tempF)
            row += 1

        # ダミー位置？
        search = "RailPos2"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        row += 1
        self.excelCell = ws.cell(row, 1)
        railNo = self.excelCell.value
        hRailNo = struct.pack("<h", railNo)
        self.newByteArr.extend(hRailNo)

        self.excelCell = ws.cell(row, 2)
        railPos = self.excelCell.value
        hRailPos = struct.pack("<h", railPos)
        self.newByteArr.extend(hRailPos)

        self.excelCell = ws.cell(row, 3)
        self.newByteArr.append(self.excelCell.value)

        self.excelCell = ws.cell(row, 4)
        f1 = self.excelCell.value
        tempF = struct.pack("<f", f1)
        self.newByteArr.extend(tempF)

        # 試運転、二人バトルの初期レール位置
        search = "FreeRunOrVSPos"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        row += 1
        for i in range(2):
            self.excelCell = ws.cell(row, 1)
            railNo = self.excelCell.value
            hRailNo = struct.pack("<h", railNo)
            self.newByteArr.extend(hRailNo)

            self.excelCell = ws.cell(row, 2)
            railPos = self.excelCell.value
            hRailPos = struct.pack("<h", railPos)
            self.newByteArr.extend(hRailPos)

            self.excelCell = ws.cell(row, 3)
            self.newByteArr.append(self.excelCell.value)

            self.excelCell = ws.cell(row, 4)
            f1 = self.excelCell.value
            tempF = struct.pack("<f", f1)
            self.newByteArr.extend(tempF)
            row += 1

        # 駅表示を始める番号
        search = "stationNo"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        self.newByteArr.append(self.excelCell.value)

        search = "RailPos4"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        row += 1
        self.excelCell = ws.cell(row, 1)
        railNo = self.excelCell.value
        hRailNo = struct.pack("<h", railNo)
        self.newByteArr.extend(hRailNo)

        self.excelCell = ws.cell(row, 2)
        railPos = self.excelCell.value
        hRailPos = struct.pack("<h", railPos)
        self.newByteArr.extend(hRailPos)

        self.excelCell = ws.cell(row, 3)
        self.newByteArr.append(self.excelCell.value)

        self.excelCell = ws.cell(row, 4)
        f1 = self.excelCell.value
        tempF = struct.pack("<f", f1)
        self.newByteArr.extend(tempF)

        search = "stationNo2"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        self.newByteArr.append(self.excelCell.value)

    def getElse1Info(self, ws):
        self.excelSheet = ws.title

        search = "else1-1"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        tempF = struct.pack("<f", self.excelCell.value)
        self.newByteArr.extend(tempF)

        search = "else1-2"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        cnt = self.excelCell.value
        self.newByteArr.append(cnt)
        row += 1
        for i in range(cnt):
            for j in range(2):
                self.excelCell = ws.cell(row, 1 + j)
                tempF = struct.pack("<f", self.excelCell.value)
                self.newByteArr.extend(tempF)
            for j in range(3):
                self.excelCell = ws.cell(row, 3 + j)
                self.newByteArr.append(self.excelCell.value)
            row += 1

        search = "light"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        lightCnt = self.excelCell.value
        self.newByteArr.append(lightCnt)
        row += 1
        for i in range(lightCnt):
            self.excelCell = ws.cell(row, 1)
            lightFile = self.excelCell.value
            bLightFile = self.encObj.convertByteArray(lightFile)
            self.newByteArr.append(len(bLightFile))
            self.newByteArr.extend(bLightFile)
            row += 1

        search = "baseBin"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        baseBinCnt = self.excelCell.value
        self.newByteArr.append(baseBinCnt)
        row += 1
        for i in range(baseBinCnt):
            self.excelCell = ws.cell(row, 1)
            baseBinFile = self.excelCell.value
            bBaseBinFile = self.encObj.convertByteArray(baseBinFile)
            self.newByteArr.append(len(bBaseBinFile))
            self.newByteArr.extend(bBaseBinFile)
            row += 1

        search = "binAnime"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        binAnimeCnt = self.excelCell.value
        self.newByteArr.append(binAnimeCnt)
        row += 1
        for i in range(binAnimeCnt):
            self.excelCell = ws.cell(row, 1)
            binIndex = self.excelCell.value
            self.newByteArr.append(binIndex)
            self.excelCell = ws.cell(row, 2)
            binAnime1 = self.excelCell.value
            hBinAnime1 = struct.pack("<h", binAnime1)
            self.newByteArr.extend(hBinAnime1)
            self.excelCell = ws.cell(row, 3)
            binAnime2 = self.excelCell.value
            hBinAnime2 = struct.pack("<h", binAnime2)
            self.newByteArr.extend(hBinAnime2)
            row += 1

    def getSmfInfo(self, ws):
        self.excelSheet = ws.title
        search = "MdlCnt"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        smfCnt = self.excelCell.value
        self.newByteArr.append(smfCnt)
        row += 1

        for i in range(smfCnt):
            self.excelCell = ws.cell(row, 2)
            smfName = self.excelCell.value
            self.newSmfList.append(smfName)

            bSmfName = self.encObj.convertByteArray(smfName)
            self.newByteArr.append(len(bSmfName))
            self.newByteArr.extend(bSmfName)

            self.excelCell = ws.cell(row, 3)
            self.newByteArr.append(self.excelCell.value)
            self.excelCell = ws.cell(row, 4)
            self.newByteArr.append(self.excelCell.value)
            self.excelCell = ws.cell(row, 5)
            self.newByteArr.append(self.excelCell.value)
            self.excelCell = ws.cell(row, 6)
            cnt = self.excelCell.value
            self.newByteArr.append(cnt)
            if cnt == 0:
                row += 1
            else:
                for j in range(cnt):
                    self.excelCell = ws.cell(row, 7)
                    tempH = struct.pack("<h", self.excelCell.value)
                    self.newByteArr.extend(tempH)
                    self.excelCell = ws.cell(row, 8)
                    tempH = struct.pack("<h", self.excelCell.value)
                    self.newByteArr.extend(tempH)
                    self.excelCell = ws.cell(row, 9)
                    tempH = struct.pack("<h", self.excelCell.value)
                    self.newByteArr.extend(tempH)
                    row += 1

    def getStationInfo(self, ws):
        self.excelSheet = ws.title

        search = "STCnt"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        stCnt = self.excelCell.value
        self.newByteArr.append(stCnt)
        row += 1
        for i in range(stCnt):
            self.excelCell = ws.cell(row, 2)
            stName = self.excelCell.value
            bStName = self.encObj.convertByteArray(stName)
            self.newByteArr.append(len(bStName))
            self.newByteArr.extend(bStName)

            self.excelCell = ws.cell(row, 3)
            stFlag = self.excelCell.value
            self.newByteArr.append(stFlag)
            self.excelCell = ws.cell(row, 4)
            railNo = self.excelCell.value
            hRailNo = struct.pack("<h", railNo)
            self.newByteArr.extend(hRailNo)
            row += 1

    def getElse2Info(self, ws):
        self.excelSheet = ws.title

        search = "else2"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        else2Cnt = self.excelCell.value
        self.newByteArr.append(else2Cnt)
        row += 1
        for i in range(else2Cnt):
            for j in range(2):
                self.excelCell = ws.cell(row, 1 + j)
                tempH = struct.pack("<h", self.excelCell.value)
                self.newByteArr.extend(tempH)
            for j in range(3):
                self.excelCell = ws.cell(row, 3 + j)
                tempF = struct.pack("<f", self.excelCell.value)
                self.newByteArr.extend(tempF)
            self.excelCell = ws.cell(row, 6)
            self.newByteArr.append(self.excelCell.value)
            row += 1

    def getCpuInfo(self, ws):
        self.excelSheet = ws.title

        search = "CPU"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        cpuCnt = self.excelCell.value
        self.newByteArr.append(cpuCnt)
        row += 1
        for i in range(cpuCnt):
            self.excelCell = ws.cell(row, 2)
            railNo = self.excelCell.value
            hRailNo = struct.pack("<h", railNo)
            self.newByteArr.extend(hRailNo)

            self.excelCell = ws.cell(row, 3)
            org = self.excelCell.value
            self.newByteArr.append(org)
            self.excelCell = ws.cell(row, 4)
            mode = self.excelCell.value
            self.newByteArr.append(mode)

            for j in range(4):
                self.excelCell = ws.cell(row, 5 + j)
                tempF = struct.pack("<f", self.excelCell.value)
                self.newByteArr.extend(tempF)
            row += 1

    def getComicScriptInfo(self, ws):
        self.excelSheet = ws.title

        search = "ComicScript"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        comicbinCnt = self.excelCell.value
        self.newByteArr.append(comicbinCnt)
        row += 1
        for i in range(comicbinCnt):
            self.excelCell = ws.cell(row, 2)
            comicNum = self.excelCell.value
            hComicNum = struct.pack("<h", comicNum)
            self.newByteArr.extend(hComicNum)

            self.excelCell = ws.cell(row, 3)
            comicType = self.excelCell.value
            self.newByteArr.append(comicType)

            self.excelCell = ws.cell(row, 4)
            railNo = self.excelCell.value
            hRailNo = struct.pack("<h", railNo)
            self.newByteArr.extend(hRailNo)
            row += 1

    def getElse4Info(self, ws):
        self.excelSheet = ws.title

        search = "else4"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        else4Cnt = self.excelCell.value
        row += 1
        for i in range(else4Cnt):
            self.excelCell = ws.cell(row, 2)
            railNo = self.excelCell.value
            if railNo not in self.else4Dict:
                self.else4Dict[railNo] = []
                for j in range(7):
                    self.excelCell = ws.cell(row, 3 + j)
                    self.else4Dict[railNo].append(self.excelCell.value)
            row += 1

    def getAmbInfo(self, ws):
        self.excelSheet = ws.title
        search = "AmbCnt"

        # AMB情報
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        ambCnt = self.excelCell.value

        row = self.findLabel("index", ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, "index"))
            return

        row += 1
        self.ambDict = {}
        for i in range(ambCnt):
            self.excelCell = ws.cell(row, 2)
            railNo = self.excelCell.value
            if railNo not in self.ambDict:
                self.ambDict[railNo] = []
            ambInfo = []
            for j in range(11):
                self.excelCell = ws.cell(row, 3 + j)
                amb = self.excelCell.value
                if j == 2:
                    amb = self.getSmfModelIndex(i, amb, self.newSmfList)
                ambInfo.append(amb)
            self.ambDict[railNo].append(ambInfo)
            row += 1

    def getElse3Info(self, ws):
        self.excelSheet = ws.title

        search = "else3"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        else3Cnt = self.excelCell.value
        row += 1
        for i in range(else3Cnt):
            self.excelCell = ws.cell(row, 2)
            railNo = self.excelCell.value
            if railNo not in self.else3Dict:
                self.else3Dict[railNo] = []
            self.excelCell = ws.cell(row, 3)
            cnt = self.excelCell.value

            for j in range(cnt):
                else3Info = []
                for k in range(5):
                    self.excelCell = ws.cell(row, 4 + k)
                    else3Info.append(self.excelCell.value)
                self.else3Dict[railNo].append(else3Info)
                row += 1

    def getRailInfo(self, ws):
        self.excelSheet = ws.title
        search = "RailCnt"
        row = self.findLabel(search, ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, search))
            return

        self.excelCell = ws.cell(row, 2)
        railCnt = self.excelCell.value

        hRailCnt = struct.pack("<h", railCnt)
        self.newByteArr.extend(hRailCnt)

        row = self.findLabel("index", ws["A"])
        if row == -1:
            self.errorLogList.append(textSetting.textList["errorList"]["E100"].format(self.excelSheet, "index"))
            return

        row += 1
        for i in range(railCnt):
            # prevRail
            self.excelCell = ws.cell(row, 2)
            prevRail = self.excelCell.value
            hPrevRail = struct.pack("<h", prevRail)
            self.newByteArr.extend(hPrevRail)

            if prevRail == -1:
                if i in self.else4Dict:
                    prevRail2 = self.else4Dict[i][0]
                    hPrevRail2 = struct.pack("<h", prevRail2)
                    self.newByteArr.extend(hPrevRail2)
                    for else4 in self.else4Dict[i][1:]:
                        tempF = struct.pack("<f", else4)
                        self.newByteArr.extend(tempF)
                else:
                    self.errorLogList.append(textSetting.textList["errorList"]["E97"].format(i))
                    return

            # block
            self.excelCell = ws.cell(row, 3)
            self.newByteArr.append(self.excelCell.value)

            # dir
            for j in range(3):
                self.excelCell = ws.cell(row, 4 + j)
                tempF = struct.pack("<f", self.excelCell.value)
                self.newByteArr.extend(tempF)

            # mdl_no
            self.excelCell = ws.cell(row, 7)
            mdl_no = self.excelCell.value
            if self.isModelNameDup(mdl_no, self.newSmfList):
                dupName = self.excelCell.value
                self.warningLogList.append(textSetting.textList["infoList"]["I115"].format(i, dupName))
            mdl_no = self.getSmfModelIndex(i, mdl_no, self.newSmfList)
            if mdl_no is None:
                return

            self.newByteArr.append(mdl_no)

            # 架線
            self.excelCell = ws.cell(row, 8)
            kasen = self.excelCell.value
            if self.isModelNameDup(kasen, self.newSmfList):
                dupName = self.excelCell.value
                self.warningLogList.append(textSetting.textList["infoList"]["I115"].format(i, dupName))
            kasen = self.getSmfModelIndex(i, kasen, self.newSmfList)
            if kasen is None:
                return

            if kasen > 127:
                bKasen = struct.pack("<B", kasen)
            else:
                bKasen = struct.pack("<b", kasen)
            self.newByteArr.extend(bKasen)

            # 架線柱
            self.excelCell = ws.cell(row, 9)
            kasenchu = self.excelCell.value
            if self.isModelNameDup(kasenchu, self.newSmfList):
                dupName = self.excelCell.value
                self.warningLogList.append(textSetting.textList["infoList"]["I115"].format(i, dupName))
            kasenchu = self.getSmfModelIndex(i, kasenchu, self.newSmfList)
            if kasenchu is None:
                return

            if kasenchu > 127:
                bKasenchu = struct.pack("<B", kasenchu)
            else:
                bKasenchu = struct.pack("<b", kasenchu)
            self.newByteArr.extend(bKasenchu)

            # per
            self.excelCell = ws.cell(row, 10)
            per = self.excelCell.value
            perF = struct.pack("<f", per)
            self.newByteArr.extend(perF)

            # flg
            for j in range(4):
                self.excelCell = ws.cell(row, 11 + j)
                flg = self.excelCell.value
                if self.flagHexMode == self.HEX_FLAG:
                    flg = int(flg, 16)
                self.newByteArr.append(flg)

            # raildata
            self.excelCell = ws.cell(row, 15)
            raildata = self.excelCell.value
            self.newByteArr.append(raildata)

            for j in range(raildata):
                self.excelCell = ws.cell(row, 16 + 4*j)
                nextRailNo = self.excelCell.value
                hNextRailNo = struct.pack("<h", nextRailNo)
                self.newByteArr.extend(hNextRailNo)
                self.excelCell = ws.cell(row, 17 + 4*j)
                nextRailPos = self.excelCell.value
                hNextRailPos = struct.pack("<h", nextRailPos)
                self.newByteArr.extend(hNextRailPos)
                self.excelCell = ws.cell(row, 18 + 4*j)
                prevRailNo = self.excelCell.value
                hPrevRailNo = struct.pack("<h", prevRailNo)
                self.newByteArr.extend(hPrevRailNo)
                self.excelCell = ws.cell(row, 19 + 4*j)
                prevRailPos = self.excelCell.value
                hPrevRailPos = struct.pack("<h", prevRailPos)
                self.newByteArr.extend(hPrevRailPos)

            # AMB情報
            if i in self.ambDict:
                ambList = self.ambDict[i]
                self.newByteArr.append(len(ambList))
                for ambInfo in ambList:
                    for j in range(4):
                        self.newByteArr.append(ambInfo[j])

                    for j in range(7):
                        tempF = struct.pack("<f", ambInfo[4 + j])
                        self.newByteArr.extend(tempF)
            else:
                self.newByteArr.append(0)
            
            # else3情報
            if i in self.else3Dict:
                else3List = self.else3Dict[i]
                self.newByteArr.append(len(else3List))
                for else3Info in else3List:
                    pos = else3Info[0]
                    self.newByteArr.append(pos)
                    railNo = else3Info[1]
                    hRailNo = struct.pack("<h", railNo)
                    self.newByteArr.extend(hRailNo)
                    binIndex = else3Info[2]
                    self.newByteArr.append(binIndex)

                    anime1 = else3Info[3]
                    hAnime1 = struct.pack("<h", anime1)
                    self.newByteArr.extend(hAnime1)
                    anime2 = else3Info[4]
                    hAnime2 = struct.pack("<h", anime2)
                    self.newByteArr.extend(hAnime2)
            else:
                self.newByteArr.append(0)
            row += 1

    def findLabel(self, search, columns):
        for column in columns:
            if column.value == search:
                return column.row
        return -1

    def getSmfModelIndex(self, i, modelValue, smfNameList):
        if type(modelValue) is str:
            if modelValue not in smfNameList:
                self.errorLogList.append(textSetting.textList["errorList"]["E96"].format(i, modelValue))
                return None
            modelIndex = smfNameList.index(modelValue)
            return modelIndex
        else:
            return modelValue

    def saveRailFile(self, filePath, newByteArr):
        w = open(filePath, "wb")
        w.write(newByteArr)
        w.close()
