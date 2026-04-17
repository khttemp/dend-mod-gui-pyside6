import os
import copy
import traceback
import openpyxl
from openpyxl.styles import PatternFill
import configparser

import program.sub.textSetting as textSetting
from program.sub.errorLogClass import ErrorLogObj


class ExcelWidget:
    def __init__(self, data, file_path, config_path, railModelInfo, ambModelInfo):
        self.errObj = ErrorLogObj()
        self.data = data
        self.filePath = file_path
        self.configPath = config_path
        self.railModelInfo = railModelInfo
        self.ambModelInfo = ambModelInfo
        self.modelNameMode = 0
        self.flagHexMode = 0
        self.ambReadMode = 0
        self.MODEL_NAME = 0
        self.HEX_FLAG = 1
        self.AMB_NEWLINE = 0
        self.row = -1
        self.errorLogList = []
        self.errorMessage = ""
        self.warningLogList = []
        self.warningMessage = ""
        self.newLinesObj = ""

        self.errorColorFill = PatternFill(patternType="solid", fgColor=textSetting.textList["excel"]["errorColor"])
        self.warningColorFill = PatternFill(patternType="solid", fgColor=textSetting.textList["excel"]["warningColor"])
        self.disableColorFill = PatternFill(patternType="solid", fgColor=textSetting.textList["excel"]["disableColor"])

    def printError(self, error):
        self.errObj.write(error)

    def extractExcel(self):
        self.errorLogList = []
        self.errorMessage = ""
        self.warningLogList = []
        self.warningMessage = ""

        wb = openpyxl.Workbook()

        defSheetNameList = wb.sheetnames
        for sheetName in defSheetNameList:
            wb.remove(wb[sheetName])

        # TabList
        tabList = textSetting.textList["ssUnity"]["ssStageDataTabList"]

        mdlList = []
        try:
            for index, tabName in enumerate(tabList):
                wb.create_sheet(index=index, title=tabName)
                self.extractStageDataInfo(self.data, index, wb[tabName], mdlList)

            if len(self.errorLogList) > 0:
                dirPath = os.path.dirname(self.filePath)
                errPath = os.path.join(dirPath, "stageError.log")
                w = open(errPath, "w", encoding="utf-8")
                for err in self.errorLogList:
                    w.write(err + "\n")
                w.close()
                self.errorMessage = textSetting.textList["errorList"]["E118"].format("stageError.log")
                return False

            wb.save(self.filePath)
            if len(self.warningLogList) > 0:
                dirPath = os.path.dirname(self.filePath)
                warnPath = os.path.join(dirPath, "stageWarning.log")
                w = open(warnPath, "w", encoding="utf-8")
                for warn in self.warningLogList:
                    w.write(warn + "\n")
                w.close()
                self.warningMessage = textSetting.textList["errorList"]["E118"].format("stageWarning.log")
            return True
        except PermissionError:
            self.errorMessage = textSetting.textList["errorList"]["E94"]
            return False
        except Exception:
            self.printError(traceback.format_exc())
            self.errorMessage = textSetting.textList["errorList"]["E14"]
            return False

    def extractStageDataInfo(self, data, sheetIndex, ws, mdlList):
        configRead = configparser.ConfigParser()
        configRead.read(self.configPath, encoding="utf-8")
        self.modelNameMode = int(configRead.get("MODEL_NAME_MODE", "mode"))
        self.flagHexMode = int(configRead.get("FLAG_MODE", "mode"))
        self.ambReadMode = int(configRead.get("AMB_READ_MODE", "mode"))

        self.row = 1
        originDataList = data.split("\n")
        # コメント行を消す
        dataList = []
        for originData in originDataList:
            if originData.find("//") == 0:
                continue
            if not originData.strip():
                continue
            dataList.append(originData.strip())

        # ストーリー、配置情報
        if sheetIndex == 0:
            self.getStoryAndPosInfo(dataList, ws)
        # 路線別画像データ
        elif sheetIndex == 1:
            self.getStageResourceInfo(dataList, ws)
        # 画像設定情報
        elif sheetIndex == 2:
            self.getTextureInfo(dataList, ws)
        # 駅名
        elif sheetIndex == 3:
            self.getStationInfo(dataList, ws)
        # ＣＰＵ切り替え
        elif sheetIndex == 4:
            self.getCpuInfo(dataList, ws)
        # コミックスクリプト
        elif sheetIndex == 5:
            self.getComicScriptInfo(dataList, ws)
        # 雨イベント
        elif sheetIndex == 6:
            self.getRainCheckerInfo(dataList, ws)
        # 土讃線スペシャル
        elif sheetIndex == 7:
            self.getDosanInfo(dataList, ws)
        # モデル情報
        elif sheetIndex == 8:
            self.getModelInfo(dataList, ws, mdlList)
        # レール情報
        elif sheetIndex == 9:
            self.getRailInfo(dataList, ws, mdlList)
        # Pri情報
        elif sheetIndex == 10:
            self.getRailPriInfo(dataList, ws)
        # AMB情報
        elif sheetIndex == 11:
            self.getAmbInfo(dataList, ws, mdlList)

    def getStoryAndPosInfo(self, dataList, ws):
        ws.cell(self.row, 1).value = "DEND_MAP_SS"
        self.row += 2

        search = "Story:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                ws.cell(self.row, 2).value = searchDataList[1]
            # Storyデータなし
            else:
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 2

        search = "Dir:"
        index = self.getSearchLine(dataList, search)
        # 必須ではない
        if index != -1:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    ws.cell(self.row, 2).value = int(searchDataList[1])
                except ValueError:
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # Dirデータなし
            else:
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 2

        search = "Track:"
        index = self.getSearchLine(dataList, search)
        # 必須ではない
        if index != -1:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    ws.cell(self.row, 2).value = int(searchDataList[1])
                except ValueError:
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # Trackデータなし
            else:
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 2

        search = "COMIC_DATA"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # COMIC_DATAの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                ws.cell(self.row, 1).value = searchDataList[0]
                # 「comic_」形式ではない
                if "comic_" not in searchDataList[0].lower():
                    ws.cell(self.row, 1).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1, searchDataList[0]))
                self.row += 1
            self.row += 1

        search = "COMIC_IMAGE"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # COMIC_IMAGEの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                ws.cell(self.row, 1).value = searchDataList[0]
                # 「comic_img_」形式ではない
                if "comic_img_" not in searchDataList[0].lower():
                    ws.cell(self.row, 1).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1, searchDataList[0]))
                self.row += 1
            self.row += 1

        search = "COMIC_SE"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # COMIC_SEの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                ws.cell(self.row, 1).value = searchDataList[0]
                # 「comic_se_」形式ではない
                if "comic_se_" not in searchDataList[0].lower():
                    ws.cell(self.row, 1).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1, searchDataList[0]))
                self.row += 1
            self.row += 1

        search = "RailPos:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # RailPosの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # railPos(rail, bone)
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1

                # railPos p
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1))
                self.row += 1
            self.row += 1

        search = "FreeRun:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            self.row += 1

            for i in range(1):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # railPos(rail, bone)
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1

                # railPos p
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1))
                self.row += 1
            self.row += 1

        search = "VSPos:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # VSPosの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # railPos(rail, bone)
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1

                # railPos p
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1))
                self.row += 1
            self.row += 1

        search = "VSStation:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    ws.cell(self.row, 2).value = int(searchDataList[1])
                except ValueError:
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # VSStationデータなし
            else:
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 2

        search = "FadeImage:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # FadeImageの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # den, name
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        ws.cell(self.row, colNum).value = val
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                self.row += 1
            self.row += 1

    def getStageResourceInfo(self, dataList, ws):
        search = "StageRes:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # StageResの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i + 1))
                idx += 1
                colNum += 1

                # den, name
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        ws.cell(self.row, colNum).value = val
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                self.row += 1
            self.row += 1

    def getTextureInfo(self, dataList, ws):
        search = "SetTexInfo:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # SetTexInfoの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # amb, amb_child, res_index
                for j in range(3):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # tex_type
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        tex_type = int(val)
                        ws.cell(self.row, colNum).value = tex_type
                    except ValueError:
                        tex_type = -1
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    tex_type = -1
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # tex_index, change_index
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                if tex_type == 31:
                    for j in range(2):
                        if realCnt > idx:
                            val = searchDataList[idx]
                            try:
                                ws.cell(self.row, colNum).value = float(val)
                            except ValueError:
                                ws.cell(self.row, colNum).value = val
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i, val))
                        else:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i))
                        idx += 1
                        colNum += 1
                self.row += 1
            self.row += 1

    def getStationInfo(self, dataList, ws):
        search = "STCnt:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # STCntの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # stIndex, rail
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # offset
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # name, jp, en
                for j in range(3):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        ws.cell(self.row, colNum).value = val
                        idx += 1
                        colNum += 1
                    else:
                        break
                self.row += 1
            self.row += 1

    def getCpuInfo(self, dataList, ws):
        search = "CPU:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # CPUの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # rail, train_no, run_type, min_len, max_len, max_speed, min_speed
                for j in range(7):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # p1(break)
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                self.row += 1
            self.row += 1

    def getComicScriptInfo(self, dataList, ws):
        search = "ComicScript:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # ComicScriptの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # event_no, event_type, rail_no
                for j in range(3):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # offset
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                self.row += 1
            self.row += 1

    def getRainCheckerInfo(self, dataList, ws):
        search = "RainChecker:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # RainCheckerの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # event_no, event_type, rail_no
                for j in range(3):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # offset
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # param
                if realCnt > idx:
                    paramCnt = len(searchDataList) - idx
                    for j in range(paramCnt):
                        if searchDataList[idx + j].find("//") == 0:
                            break
                        val = searchDataList[idx + j]
                        try:
                            ws.cell(self.row, colNum + j).value = float(val)
                        except ValueError:
                            ws.cell(self.row, colNum + j).value = val
                            ws.cell(self.row, colNum + j).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                self.row += 1
            self.row += 1

    def getDosanInfo(self, dataList, ws):
        search = "DosanInfo:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # DosanInfoの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # event_no, event_type, rail_no
                for j in range(3):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # offset
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # param
                if realCnt > idx:
                    paramCnt = len(searchDataList) - idx
                    for j in range(paramCnt):
                        val = searchDataList[idx + j]
                        try:
                            ws.cell(self.row, colNum + j).value = float(val)
                        except ValueError:
                            ws.cell(self.row, colNum + j).value = val
                            ws.cell(self.row, colNum + j).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                self.row += 1
            self.row += 1

    def getModelInfo(self, dataList, ws, mdlList):
        search = "MdlCnt:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # MdlCntの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # mdl_name
                if realCnt > idx:
                    val = searchDataList[idx]
                    ws.cell(self.row, colNum).value = val
                    mdlList.append(val)
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # flg, flg
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            flg = int(val)
                            if self.flagHexMode == self.HEX_FLAG:
                                flg = self.toHex(flg)
                            ws.cell(self.row, colNum).value = flg
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # kasenchu_mdl
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                self.row += 1
            self.row += 1

    def getRailInfo(self, dataList, ws, mdlList):
        search = "RailCnt:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # RailCntの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 2

            titleList = [
                "index",
                "prev_rail",
                "block",
                "pos_x",
                "pos_y",
                "pos_z",
                "dir_x",
                "dir_y",
                "dir_z",
                "mdl_no",
                "mdl_kasenchu",
                "per",
                "flg",
                "flg",
                "flg",
                "flg",
                "rail_data",
                "next_rail",
                "next_no",
                "prev_rail",
                "prev_no",
            ]

            for idx, title in enumerate(titleList):
                ws.cell(self.row, 1 + idx).value = title
                idx += 1
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                real_mdl_name = None
                disableFlag = False
                if realCnt > 15:
                    try:
                        railFlg = int(searchDataList[15])
                        if railFlg & 0x80 != 0:
                            disableFlag = True
                    except ValueError:
                        disableFlag = False

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    if not disableFlag:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                if disableFlag:
                    ws.cell(self.row, colNum).fill = self.disableColorFill
                idx += 1
                colNum += 1

                # prev_rail, block
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            if not disableFlag:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i))
                    if disableFlag:
                        ws.cell(self.row, colNum).fill = self.disableColorFill
                    idx += 1
                    colNum += 1

                # pos, dir (xyz)
                for j in range(6):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = float(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            if not disableFlag:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i))
                    if disableFlag:
                        ws.cell(self.row, colNum).fill = self.disableColorFill
                    idx += 1
                    colNum += 1

                # mdl_no
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        mdl_no = int(val)
                        if self.modelNameMode == self.MODEL_NAME:
                            mdl_name = self.getModelName(mdl_no, mdlList)
                            real_mdl_name = self.getModelName(mdl_no, mdlList, False)
                            if not str(real_mdl_name).isdigit():
                                if real_mdl_name.lower() not in list(self.railModelInfo.keys()):
                                    if not disableFlag:
                                        ws.cell(self.row, colNum).fill = self.errorColorFill
                                        self.errorLogList.append(self.notAvailableRail(i, real_mdl_name))
                                    real_mdl_name = None
                            else:
                                real_mdl_name = None
                        else:
                            mdl_name = mdl_no
                        ws.cell(self.row, colNum).value = mdl_name
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    if not disableFlag:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                if disableFlag:
                    ws.cell(self.row, colNum).fill = self.disableColorFill
                idx += 1
                colNum += 1

                # mdl_kasenchu
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        mdl_kasenchu = int(val)
                        if self.modelNameMode == self.MODEL_NAME:
                            mdl_name = self.getModelName(mdl_kasenchu, mdlList)
                        else:
                            mdl_name = mdl_kasenchu
                        ws.cell(self.row, colNum).value = mdl_name
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    if not disableFlag:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                if disableFlag:
                    ws.cell(self.row, colNum).fill = self.disableColorFill
                idx += 1
                colNum += 1

                # per
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(self.row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    if not disableFlag:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                if disableFlag:
                    ws.cell(self.row, colNum).fill = self.disableColorFill
                idx += 1
                colNum += 1

                # flg, flg, flg, flg
                for j in range(4):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            flg = int(val)
                            if self.flagHexMode == self.HEX_FLAG:
                                flg = self.toHex(flg)
                            ws.cell(self.row, colNum).value = flg
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            if not disableFlag:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i))
                    if disableFlag:
                        ws.cell(self.row, colNum).fill = self.disableColorFill
                    idx += 1
                    colNum += 1

                # rail_data
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        rail_data = int(val)
                        if real_mdl_name is not None:
                            real_rail_data = self.railModelInfo[real_mdl_name.lower()]
                            if rail_data != real_rail_data:
                                if not disableFlag:
                                    ws.cell(self.row, colNum).fill = self.warningColorFill
                                    self.warningLogList.append(self.diffRailDataError(i, real_mdl_name, real_rail_data, rail_data))
                        ws.cell(self.row, colNum).value = rail_data
                    except ValueError:
                        rail_data = 0
                        ws.cell(self.row, colNum).value = val
                        if not disableFlag:
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                else:
                    rail_data = 0
                    if not disableFlag:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                if disableFlag:
                    ws.cell(self.row, colNum).fill = self.disableColorFill
                idx += 1
                colNum += 1

                for j in range(rail_data):
                    for k in range(4):
                        if realCnt > idx:
                            val = searchDataList[idx]
                            try:
                                ws.cell(self.row, colNum).value = int(val)
                            except ValueError:
                                ws.cell(self.row, colNum).value = val
                                if not disableFlag:
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i, val))
                        else:
                            if not disableFlag:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i))
                        if disableFlag:
                            ws.cell(self.row, colNum).fill = self.disableColorFill
                        idx += 1
                        colNum += 1
                self.row += 1
            self.row += 1

    def getRailPriInfo(self, dataList, ws):
        search = "RailPri:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # RailPriの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                self.row += 1
            self.row += 1

        search = "BtlPri:"
        index = self.getSearchLine(dataList, search)
        # 必須ではない
        if index != -1:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # BtlPriの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                self.row += 1
            self.row += 1

        search = "NoDriftRail:"
        index = self.getSearchLine(dataList, search)
        # 必須ではない
        if index != -1:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # NoDriftRailの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))
            self.row += 1

            for i in range(cnt):
                searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i + 1, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                self.row += 1
            self.row += 1

    def getAmbInfo(self, dataList, ws, mdlList):
        search = "AmbCnt:"
        index = self.getSearchLine(dataList, search)
        if index == -1:
            self.errorLogList.append(self.failSearchError(search))
        else:
            searchDataList = self.getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(self.row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(self.row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(self.row, 2).value = searchDataList[1]
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[1]))
            # AmbCntの数なし
            else:
                cnt = 0
                ws.cell(self.row, 2).fill = self.errorColorFill
                self.errorLogList.append(self.noCntDataError(search))

            if len(searchDataList) > 2:
                try:
                    sizeFlag = int(searchDataList[2])
                    ws.cell(self.row, 3).value = sizeFlag
                except ValueError:
                    sizeFlag = 0
                    ws.cell(self.row, 3).value = searchDataList[2]
                    ws.cell(self.row, 3).fill = self.errorColorFill
                    self.errorLogList.append(self.cntDataReadError(search, searchDataList[2]))
            else:
                sizeFlag = 0
            self.row += 2

            titleList = [
                "index",
                "rail",
                "length",
                "amd_data",
                "mdl_no",
                "parentIndex",
                "pos_x",
                "pos_y",
                "pos_z",
                "dir_x",
                "dir_y",
                "dir_z",
                "joint_dir_x",
                "joint_dir_y",
                "joint_dir_z",
                "per",
                "kasenchu_per",
            ]

            for idx, title in enumerate(titleList):
                ws.cell(self.row, 1 + idx).value = title
                idx += 1
            self.row += 1

            for i in range(cnt):
                try:
                    searchDataList = self.getSplitAndRemoveEmptyData(dataList[index + i + 1])
                except IndexError:
                    self.errorLogList.append(self.outOfRangeError(search, cnt, i))
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                # index
                if realCnt > idx:
                    val = i
                    ws.cell(self.row, colNum).value = val
                else:
                    ws.cell(self.row, colNum).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                idx += 1
                colNum += 1

                # rail_no, length
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(self.row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(self.row, colNum).value = val
                            ws.cell(self.row, colNum).fill = self.errorColorFill
                            self.errorLogList.append(self.dataReadError(search, i, val))
                    else:
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i))
                    idx += 1
                    colNum += 1

                # amb_data
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ambData = int(val)
                        ws.cell(self.row, colNum).value = ambData
                    except ValueError:
                        ws.cell(self.row, colNum).value = val
                        ws.cell(self.row, colNum).fill = self.errorColorFill
                        self.errorLogList.append(self.dataReadError(search, i, val))
                        continue
                else:
                    ambData = None
                idx += 1
                colNum += 1

                if ambData is None:
                    ws.cell(self.row, 1).fill = self.errorColorFill
                    ws.cell(self.row, 2).fill = self.errorColorFill
                    ws.cell(self.row, 3).fill = self.errorColorFill
                    ws.cell(self.row, 4).fill = self.errorColorFill
                    self.errorLogList.append(self.dataReadError(search, i))
                    self.row += 1
                else:
                    if ambData <= 0:
                        ws.cell(self.row, 1).fill = self.warningColorFill
                        ws.cell(self.row, 2).fill = self.warningColorFill
                        ws.cell(self.row, 3).fill = self.warningColorFill
                        ws.cell(self.row, 4).fill = self.warningColorFill
                        self.warningLogList.append(self.ambDataWarning(i))
                        # ambDataは0個なのに、データがある
                        if len(searchDataList) > idx:
                            for j in range(idx, len(searchDataList)):
                                try:
                                    ws.cell(self.row, j + 1).value = float(searchDataList[j])
                                    ws.cell(self.row, j + 1).fill = self.warningColorFill
                                except ValueError:
                                    ws.cell(self.row, j + 1).value = searchDataList[j]
                                    ws.cell(self.row, j + 1).fill = self.errorColorFill
                        self.row += 1
                    else:
                        for j in range(ambData):
                            if self.ambReadMode == self.AMB_NEWLINE:
                                colNum = 5
                            # mdl_no
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    mdl_no = int(val)
                                    if self.modelNameMode == self.MODEL_NAME:
                                        mdl_name = self.getModelName(mdl_no, mdlList)
                                        real_mdl_name = self.getModelName(mdl_no, mdlList, False)
                                        if not str(real_mdl_name).isdigit():
                                            if real_mdl_name.lower() not in self.ambModelInfo:
                                                ws.cell(self.row, colNum).fill = self.warningColorFill
                                                self.warningLogList.append(self.notAvailableAmb(i, real_mdl_name))
                                    else:
                                        mdl_name = mdl_no
                                    ws.cell(self.row, colNum).value = mdl_name
                                except ValueError:
                                    ws.cell(self.row, colNum).value = val
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i, val))
                            else:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i))
                            idx += 1
                            colNum += 1

                            # parentindex
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    ws.cell(self.row, colNum).value = int(val)
                                except ValueError:
                                    ws.cell(self.row, colNum).value = val
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i, val))
                            else:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i))
                            idx += 1
                            colNum += 1

                            # pos, dir, joint_dir (xyz)
                            for j in range(9):
                                if realCnt > idx:
                                    val = searchDataList[idx]
                                    try:
                                        ws.cell(self.row, colNum).value = float(val)
                                    except ValueError:
                                        ws.cell(self.row, colNum).value = val
                                        ws.cell(self.row, colNum).fill = self.errorColorFill
                                        self.errorLogList.append(self.dataReadError(search, i, val))
                                else:
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i))
                                idx += 1
                                colNum += 1

                            # per
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    ws.cell(self.row, colNum).value = float(val)
                                except ValueError:
                                    ws.cell(self.row, colNum).value = val
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i, val))
                            else:
                                ws.cell(self.row, colNum).fill = self.errorColorFill
                                self.errorLogList.append(self.dataReadError(search, i))
                            idx += 1
                            colNum += 1

                            # size_per
                            if sizeFlag >= 1:
                                if realCnt > idx:
                                    val = searchDataList[idx]
                                    try:
                                        ws.cell(self.row, colNum).value = float(val)
                                    except ValueError:
                                        ws.cell(self.row, colNum).value = val
                                        ws.cell(self.row, colNum).fill = self.errorColorFill
                                        self.errorLogList.append(self.dataReadError(search, i, val))
                                else:
                                    ws.cell(self.row, colNum).fill = self.errorColorFill
                                    self.errorLogList.append(self.dataReadError(search, i))
                                idx += 1
                                colNum += 1
                            if self.ambReadMode == self.AMB_NEWLINE:
                                self.row += 1
                        moreFlag = False
                        if self.ambReadMode == self.AMB_NEWLINE:
                            moreRow = self.row - 1
                        else:
                            moreRow = self.row
                        # データがもっとある
                        while realCnt > idx + 12:
                            moreFlag = True
                            # mdl_no
                            val = searchDataList[idx]
                            try:
                                mdl_no = int(val)
                                if self.modelNameMode == self.MODEL_NAME:
                                    mdl_name = self.getModelName(mdl_no, mdlList)
                                    real_mdl_name = self.getModelName(mdl_no, mdlList, False)
                                    if not str(real_mdl_name).isdigit():
                                        if real_mdl_name.lower() not in self.ambModelInfo:
                                            ws.cell(self.row, colNum).fill = self.warningColorFill
                                            self.warningLogList.append(self.notAvailableAmb(i, real_mdl_name))
                                else:
                                    mdl_name = mdl_no
                                ws.cell(moreRow, colNum).value = mdl_name
                                ws.cell(moreRow, colNum).fill = self.warningColorFill
                            except ValueError:
                                ws.cell(moreRow, colNum).value = val
                                ws.cell(moreRow, colNum).fill = self.errorColorFill
                            idx += 1
                            colNum += 1

                            # parentindex
                            val = searchDataList[idx]
                            try:
                                ws.cell(moreRow, colNum).value = int(val)
                                ws.cell(moreRow, colNum).fill = self.warningColorFill
                            except ValueError:
                                ws.cell(moreRow, colNum).value = val
                                ws.cell(moreRow, colNum).fill = self.errorColorFill
                            idx += 1
                            colNum += 1

                            # pos, dir, joint_dir (xyz)
                            for j in range(9):
                                val = searchDataList[idx]
                                try:
                                    ws.cell(moreRow, colNum).value = float(val)
                                    ws.cell(moreRow, colNum).fill = self.warningColorFill
                                except ValueError:
                                    ws.cell(moreRow, colNum).value = val
                                    ws.cell(moreRow, colNum).fill = self.errorColorFill
                                idx += 1
                                colNum += 1

                            # per
                            val = searchDataList[idx]
                            try:
                                ws.cell(moreRow, colNum).value = float(val)
                                ws.cell(moreRow, colNum).fill = self.warningColorFill
                            except ValueError:
                                ws.cell(moreRow, colNum).value = val
                                ws.cell(moreRow, colNum).fill = self.errorColorFill
                            idx += 1
                            colNum += 1

                            # size_per
                            if sizeFlag >= 1:
                                if realCnt > idx:
                                    val = searchDataList[idx]
                                    try:
                                        ws.cell(moreRow, colNum).value = float(val)
                                        ws.cell(moreRow, colNum).fill = self.warningColorFill
                                    except ValueError:
                                        ws.cell(moreRow, colNum).value = val
                                        ws.cell(moreRow, colNum).fill = self.errorColorFill
                                    idx += 1
                                    colNum += 1
                        if self.ambReadMode != self.AMB_NEWLINE:
                            self.row += 1
                        if moreFlag:
                            self.warningLogList.append(self.notUsedAmbData(i))
            self.row += 1

    def getSearchLine(self, dataList, search):
        for idx, data in enumerate(dataList):
            if data.find(search) == 0:
                return idx
        return -1

    def getSplitAndRemoveEmptyData(self, data):
        sList = data.strip().split("\t")
        sList = list(filter(None, sList))
        sList = [x for x in sList if x.find("//") != 0]
        return sList

    def getModelName(self, mdl_no, mdlList, dupFlag=True):
        if mdl_no < 0 or mdl_no >= len(mdlList):
            return mdl_no
        mdlName = mdlList[mdl_no]
        if dupFlag:
            dupList = [x for x in mdlList if x == mdlName]
            if len(dupList) > 1:
                return mdl_no
            else:
                return mdlName
        else:
            return mdlName

    def toHex(self, num):
        return "0x{:02x}".format(num)

    def failSearchError(self, search):
        return textSetting.textList["errorList"]["E106"].format(search)

    def outOfRangeError(self, search, cnt, i):
        return textSetting.textList["errorList"]["E112"].format(search, cnt, i)

    def ambDataWarning(self, i):
        return textSetting.textList["errorList"]["E113"].format(i)

    def dataReadError(self, search, i, data="（データなし）"):
        return textSetting.textList["errorList"]["E114"].format(search, i, data)
    
    def noCntDataError(self, search):
        return textSetting.textList["errorList"]["E116"].format(search)

    def cntDataReadError(self, search, data):
        return textSetting.textList["errorList"]["E117"].format(search, data)

    def notUsedAmbData(self, i):
        return textSetting.textList["errorList"]["E119"].format(i)

    def notAvailableAmb(self, i, model_name):
        return textSetting.textList["errorList"]["E120"].format(i, model_name)

    def notAvailableRail(self, i, model_name):
        return textSetting.textList["errorList"]["E121"].format(i, model_name)

    def diffRailDataError(self, i, model_name, realCnt, cnt):
        if realCnt == 1:
            realRail = "単線"
        else:
            realRail = "複線"

        if cnt == 1:
            curRail = "単線"
        else:
            curRail = "複線"
        return textSetting.textList["errorList"]["E122"].format(i, model_name, realRail, curRail)

    def loadExcelAndMerge(self):
        self.errorMessage = ""
        self.warningMessage = ""
        self.errorLogList = []
        self.warningLogList = []
        self.newLinesObj = ""

        originData = self.data
        newLines = copy.deepcopy(originData).split("\n")
        wb = openpyxl.load_workbook(self.filePath, data_only=True)
        tabList = textSetting.textList["ssUnity"]["ssStageDataTabList"]

        oneSheetFlag = False
        for tabName in tabList:
            if tabName not in wb.sheetnames:
                oneSheetFlag = True
                break

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("Story:", readSheet, newLines, dataFlag=False)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("Dir:", readSheet, newLines, dataFlag=False, requiredFlag=False)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("Track:", readSheet, newLines, dataFlag=False, requiredFlag=False, otherSearchList=["Dir:", "Story:"])

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("COMIC_DATA", readSheet, newLines)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("COMIC_IMAGE", readSheet, newLines)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("COMIC_SE", readSheet, newLines)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("RailPos:", readSheet, newLines, optionalRead=0)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("FreeRun:", readSheet, newLines, headerDataFlag=False, optionalRead=0)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("VSPos:", readSheet, newLines, optionalRead=0)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("VSStation:", readSheet, newLines, dataFlag=False)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[0]]
        self.findSearchAndSetCnt("FadeImage:", readSheet, newLines, optionalRead=1)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[1]]
        self.findSearchAndSetCnt("StageRes:", readSheet, newLines, optionalRead=0)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[2]]
        self.findSearchAndSetCnt("SetTexInfo:", readSheet, newLines, optionalRead=2)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[3]]
        self.findSearchAndSetCnt("STCnt:", readSheet, newLines, optionalRead=3)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[4]]
        self.findSearchAndSetCnt("CPU:", readSheet, newLines, optionalRead=4)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[5]]
        self.findSearchAndSetCnt("ComicScript:", readSheet, newLines, optionalRead=5)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[6]]
        self.findSearchAndSetCnt("RainChecker:", readSheet, newLines, optionalRead=6)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[7]]
        self.findSearchAndSetCnt("DosanInfo:", readSheet, newLines, optionalRead=6)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[8]]
        newMdlList = []
        self.findSearchAndSetCnt("MdlCnt:", readSheet, newLines, optionalRead=7, mdlList=newMdlList)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[9]]
        self.findSearchAndSetCnt("RailCnt:", readSheet, newLines, optionalRead=8, mdlList=newMdlList)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[10]]
        self.findSearchAndSetCnt("RailPri:", readSheet, newLines, optionalRead=1)

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[10]]
        self.findSearchAndSetCnt("BtlPri:", readSheet, newLines, requiredFlag=False, optionalRead=1, otherSearchList=["RailPri:"])

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[10]]
        self.findSearchAndSetCnt("NoDriftRail:", readSheet, newLines, requiredFlag=False, optionalRead=1, otherSearchList=["BtlPri:", "RailPri:"])

        readSheet = wb.worksheets[0] if oneSheetFlag else wb[tabList[11]]
        self.findSearchAndSetCnt("AmbCnt:", readSheet, newLines, optionalRead=9, mdlList=newMdlList)
        
        if len(self.errorLogList) > 0:
            dirPath = os.path.dirname(self.filePath)
            errPath = os.path.join(dirPath, "stageError.log")
            w = open(errPath, "w", encoding="utf-8")
            for err in self.errorLogList:
                w.write(err + "\n")
            w.close()
            self.errorMessage = textSetting.textList["errorList"]["E118"].format("stageError.log")
            return False

        if len(self.warningLogList) > 0:
            warnPath = os.path.join(dirPath, "stageWarning.log")
            w = open(warnPath, "w", encoding="utf-8")
            for warn in self.warningLogList:
                w.write(warn + "\n")
            w.close()
            self.warningMessage = textSetting.textList["errorList"]["E118"].format("stageWarning.log")
        self.newLinesObj = newLines
        return True

    def findSearchAndSetCnt(self, search, ws, newLines, headerDataFlag=True, dataFlag=True, requiredFlag=True, optionalRead=-1, otherSearchList=["Story:"], mdlList=[]):
        configRead = configparser.ConfigParser()
        configRead.read(self.configPath, encoding="utf-8")
        self.flagHexMode = int(configRead.get("FLAG_MODE", "mode"))
        self.ambReadMode = int(configRead.get("AMB_READ_MODE", "mode"))

        eIndex = self.findLabel(search, ws["A"])
        # エクセルでラベルを見つけられない
        if eIndex == -1:
            # 必須項目の場合、エラー
            if requiredFlag:
                self.errorLogList.append(self.failExcelSearchError(search))
                return
            # 必須項目ではない場合
            else:
                index = self.getSearchLine(newLines, search)
                # denで見つけた場合、消す
                if index != -1:
                    if dataFlag:
                        delcnt = int(self.getSplitAndRemoveEmptyData(newLines[index])[1])
                    else:
                        delcnt = 1
                    index -= 1
                    while newLines[index].find("//") == 0:
                        newLines.pop(index)
                        index -= 1
                    index += 1
                    newLines.pop(index)
                    for i in range(delcnt):
                        newLines.pop(index)
                return
        # エクセルでラベルを見つけた場合、ラベルがある行全体を取得
        valList = []
        for cell in ws[eIndex]:
            if cell.value is None:
                break
            valList.append(str(cell.value))
        # データ数が必要な項目で、2個より少ない場合
        if headerDataFlag:
            if len(valList) < 2:
                self.errorLogList.append(self.failExcelValueError(search))
                return 
            newLine = "\t".join(valList)
            try:
                newCnt = int(valList[1])
            except ValueError:
                newCnt = 1
        # データ数が必要ない、または1個のみ
        else:
            newLine = "".join(valList)
            newCnt = 1
        newLine += "\r"

        # denで、ラベルを探す
        index = self.getSearchLine(newLines, search)
        if index == -1:
            # 必須項目の場合、エラー
            if requiredFlag:
                self.errorLogList.append(self.failSearchError(search))
                return
            # 必須項目ではないが、denのデータにない場合、追加
            else:
                for otherSearch in otherSearchList:
                    index = self.getSearchLine(newLines, otherSearch)
                    if index != -1:
                        break
                if dataFlag:
                    cnt = int(self.getSplitAndRemoveEmptyData(newLines[index])[1])
                    index += cnt
                index += 1
                newLines.insert(index, newLine)
                newLines.insert(index, "\r")

        # データ数が必要な項目の場合
        if headerDataFlag:
            # denでラベルを探す
            index = self.getSearchLine(newLines, search)
            # 元のデータ数を取得
            try:
                originCnt = int(self.getSplitAndRemoveEmptyData(newLines[index])[1])
            except ValueError:
                originCnt = 1
        else:
            originCnt = 1
        # denにエクセルのラベルで更新
        newLines[index] = newLine

        # データ数通り、読み込む
        if dataFlag:
            newDataList = []

            startIndex = eIndex + 1
            if newCnt > 0:
                while True:
                    val = ws.cell(startIndex, 1).value
                    if val is not None and str(val).find("//") != 0:
                        break
                    startIndex += 1

                # レールデータやAMB
                if optionalRead == 8 or optionalRead == 9:
                    val = ws.cell(startIndex, 1).value
                    if "index" in str(val):
                        startIndex += 1

            for i in range(newCnt):
                # デフォルト読み(1個読み)
                if optionalRead == -1:
                    val = ws.cell(startIndex + i, 1).value
                    # 読み込むデータが空白の場合、エラー
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1).coordinate
                        self.errorLogList.append(self.failExcelValueError(search, coordinate))
                        return
                    newDataList.append("{0}\r".format(val))
                # データ３つよみ
                elif optionalRead == 0:
                    columnList = []
                    for j in range(3):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # データ２つ読み
                elif optionalRead == 1:
                    columnList = []
                    for j in range(2):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # SetTexInfoの読み
                elif optionalRead == 2:
                    columnList = []
                    tex_type = -1
                    for j in range(7):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                        if j == 4:
                            tex_type = int(val)
                    val8 = ws.cell(startIndex + i, 8).value
                    val9 = ws.cell(startIndex + i, 9).value
                    if tex_type == 31 and val8 is not None and val9 is not None:
                        columnList.append("{0}".format(val8))
                        columnList.append("{0}".format(val9))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # 駅名の読み
                elif optionalRead == 3:
                    columnList = []
                    for j in range(4):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    val5 = ws.cell(startIndex + i, 5).value
                    val6 = ws.cell(startIndex + i, 6).value
                    val7 = ws.cell(startIndex + i, 7).value
                    if val5 is not None:
                        columnList.append("{0}".format(val5))
                    if val6 is not None:
                        columnList.append("{0}".format(val6))
                    if val7 is not None:
                        columnList.append("{0}".format(val7))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # ＣＰＵの読み
                elif optionalRead == 4:
                    columnList = []
                    for j in range(8):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    val9 = ws.cell(startIndex + i, 9).value
                    if val9 is not None:
                        columnList.append("{0}".format(val9))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # データ５つの読み
                elif optionalRead == 5:
                    columnList = []
                    for j in range(5):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # イベントの読み
                elif optionalRead == 6:
                    columnList = []
                    for j in range(5):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    colIdx = 6
                    while True:
                        val = ws.cell(startIndex + i, colIdx).value
                        if val is None:
                            break
                        columnList.append("{0}".format(val))
                        colIdx += 1
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # モデルの読み
                elif optionalRead == 7:
                    columnList = []
                    for j in range(5):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        if j == 1:
                            mdlList.append(val)
                        if j in [2, 3]:
                            if self.flagHexMode == self.HEX_FLAG:
                                try:
                                    val = int(val, 16)
                                except TypeError:
                                    originValue = ws.cell(startIndex + i, 1 + j).value
                                    coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                                    self.errorLogList.append(self.failHexValueError(search, coordinate, originValue))
                                    return
                            else:
                                if type(val) == str and "0x" in val:
                                    originValue = ws.cell(startIndex + i, 1 + j).value
                                    coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                                    self.errorLogList.append(self.notConvertHexValueError(search, coordinate, originValue))
                                    return
                        columnList.append("{0}".format(val))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # レールデータの読み
                elif optionalRead == 8:
                    columnList = []
                    for j in range(16):
                        val = ws.cell(startIndex + i, 1 + j).value
                        # 読み込むデータが空白の場合、エラー
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        # mdl_no
                        if j == 9:
                            val = self.getModelIndex(val, startIndex + i, mdlList, search)
                            if val is None:
                                return
                        # mdl_kasenchu
                        elif j == 10:
                            val = self.getModelIndex(val, startIndex + i, mdlList, search)
                            if val is None:
                                return
                        # flg
                        elif j >= 12 and j <= 15:
                            if self.flagHexMode == self.HEX_FLAG:
                                try:
                                    val = int(val, 16)
                                except TypeError:
                                    originValue = ws.cell(startIndex + i, 1 + j).value
                                    coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                                    self.errorLogList.append(self.failHexValueError(search, coordinate, originValue))
                                    return
                            else:
                                if type(val) == str and "0x" in val:
                                    originValue = ws.cell(startIndex + i, 1 + j).value
                                    coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                                    self.errorLogList.append(self.notConvertHexValueError(search, coordinate, originValue))
                                    return
                        columnList.append("{0}".format(val))
                    rail_data = ws.cell(startIndex + i, 17).value
                    # 単線複線の設定が空白の場合、エラー
                    if rail_data is None:
                        coordinate = ws.cell(startIndex + i, 17).coordinate
                        self.errorLogList.append(self.failExcelValueError(search, coordinate))
                        return
                    columnList.append("{0}".format(rail_data))

                    for j in range(rail_data):
                        for k in range(4):
                            val = ws.cell(startIndex + i, 18 + 4*j + k).value
                            # next, prevの設定が空の場合、エラー
                            if val is None:
                                coordinate = ws.cell(startIndex + i, 18 + 4*j + k).coordinate
                                self.errorLogList.append(self.failExcelValueError(search, coordinate))
                                return
                            columnList.append("{0}".format(val))
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))
                # AMBデータの読み
                elif optionalRead == 9:
                    columnList = []
                    for j in range(3):
                        val = ws.cell(startIndex, 1 + j).value
                        if val is None:
                            coordinate = ws.cell(startIndex, 1 + j).coordinate
                            self.errorLogList.append(self.failExcelValueError(search, coordinate))
                            return
                        columnList.append("{0}".format(val))
                    amb_data = ws.cell(startIndex, 4).value
                    columnList.append("{0}".format(amb_data))

                    if amb_data <= 0:
                        startIndex += 1
                    else:
                        for j in range(amb_data):
                            if self.ambReadMode == self.AMB_NEWLINE:
                                columnStart = 5
                            else:
                                columnStart = 5 + 13*j
                            for k in range(13):
                                val = ws.cell(startIndex, columnStart + k).value
                                # 読み込むデータが空白の場合、エラー
                                if val is None:
                                    coordinate = ws.cell(startIndex, columnStart + k).coordinate
                                    self.errorLogList.append(self.failExcelValueError(search, coordinate))
                                    return
                                # mdl_no
                                if k == 0:
                                    val = self.getModelIndex(val, startIndex, mdlList, search)
                                    if val is None:
                                        return
                                columnList.append("{0}".format(val))
                            if self.ambReadMode == self.AMB_NEWLINE:
                                startIndex += 1
                        if self.ambReadMode != self.AMB_NEWLINE:
                            startIndex += 1
                    columnLine = "\t".join(columnList)
                    newDataList.append("{0}\r".format(columnLine))

            newDataList.reverse()

            index += 1
            delCnt = 0
            while delCnt < originCnt:
                if not newLines[index].strip():
                    delCnt -= 1
                newLines.pop(index)
                delCnt += 1
            for i in range(newCnt):
                newLines.insert(index, newDataList[i])
        return

    def findLabel(self, search, columns):
        for column in columns:
            if column.value == search:
                return column.row
        return -1

    def failExcelSearchError(self, search):
        return textSetting.textList["errorList"]["E107"].format(search)

    def failExcelValueError(self, search, coord=None):
        if coord is not None:
            return textSetting.textList["errorList"]["E111"].format(search, coord)
        return textSetting.textList["errorList"]["E108"].format(search)

    def failHexValueError(self, search, coord, val):
        return textSetting.textList["errorList"]["E129"].format(search, coord, val)

    def notConvertHexValueError(self, search, coord, val):
        return textSetting.textList["errorList"]["E130"].format(search, coord, val)

    def getModelIndex(self, val, idx, mdlList, search):
        if type(val) is str:
            # 存在しないモデル名称の場合、エラー
            if val not in mdlList:
                if search == "RailCnt:":
                    self.errorLogList.append(textSetting.textList["errorList"]["E109"].format(val))
                elif search == "AmbCnt:":
                    self.errorLogList.append(textSetting.textList["errorList"]["E110"].format(val))
                return None
            tempList = [x for x in mdlList if x == val]
            # MdlCntに重複設定しているものがある場合、Warning
            if len(tempList) > 1:
                if search == "RailCnt:":
                    self.warningLogList.append(textSetting.textList["infoList"]["I115"].format(idx, val))
                elif search == "AmbCnt:":
                    self.warningLogList.append(textSetting.textList["infoList"]["I116"].format(idx, val))
            modelIndex = mdlList.index(val)
            return modelIndex
        else:
            return val
